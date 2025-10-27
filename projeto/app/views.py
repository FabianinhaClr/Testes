import os
import re
import time
import requests
import pandas as pd
import calendar
from datetime import date, datetime
from dateutil import parser as dparser
from io import BytesIO
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import load_workbook, Workbook
from dotenv import load_dotenv


load_dotenv()

API_URL = os.getenv("API_URL")
API_KEY = os.getenv("API_KEY")
SLEEP = 0.5
DEBUG = True


# ---------------- View Django com download ---------------- #
@login_required(login_url="login")   # <<--- protege com login
def upload_file(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]
        try:
            # (SEU CÓDIGO EXISTENTE AQUI, SEM MUDAR)
            # ...
            return response  # seu HttpResponse do Excel
        except Exception as e:
            return render(request, "upload.html", {"mensagem": f"❌ Erro ao processar: {e}"})
    return render(request, "upload.html")


# ------- ADICIONE ESTAS VIEWS DE AUTENTICAÇÃO ------- #
def login_view(request):
    if request.method == "POST":
        user = authenticate(
            request,
            username=request.POST.get("username"),
            password=request.POST.get("password"),
        )
        if user:
            login(request, user)
            return redirect("upload")  # vai para a tela de upload
        return render(request, "login.html", {"error": "Usuário ou senha inválidos."})
    return render(request, "login.html")

def logout_view(request):
    logout(request)
    return redirect("login")

# ---------------- Funções do script original ---------------- #
def clean_cnpj(s):
    return re.sub(r'\D', '', str(s)).zfill(14)

def read_cnpjs(file):
    # Sempre lê a primeira aba para pegar os CNPJs
    df = pd.read_excel(file, sheet_name=0)
    if 'cnpj_part' not in df.columns:
        raise ValueError("A planilha deve conter a coluna 'cnpj_part'")
    cnpjs = df['cnpj_part'].dropna().astype(str).apply(lambda x: re.sub(r'\D', '', x).zfill(14))
    return cnpjs.unique()

def query_infosimples(cnpj):
    try:
        if API_URL and API_URL.strip():
            args = {"cnpj": cnpj, "token": API_KEY, "timeout": 300}
            r = requests.post(API_URL, data=args, timeout=60)
        else:
            r = requests.get(f"https://www.receitaws.com.br/v1/cnpj/{cnpj}", timeout=60)
        try:
            j = r.json()
        except Exception:
            j = None
        if DEBUG:
            print("="*80)
            print(f"[DEBUG] CNPJ PROCESSADO: {cnpj}")
            import json
            print(json.dumps(j, indent=2, ensure_ascii=False))
        return r.status_code, j
    except Exception as e:
        if DEBUG:
            print(f"[DEBUG] Erro na requisição para {cnpj}: {e}")
        return None, None

def parse_date_any(s):
    if not s:
        return None
    s = str(s).strip()
    formatos = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"]
    for fmt in formatos:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    try:
        return dparser.parse(s, dayfirst=True).date()
    except Exception:
        return None

def _get_value(item, keys):
    for k in keys:
        if isinstance(item, dict) and k in item and item[k] not in (None, ""):
            return item[k]
    return None

def extract_periods_from_response(resp_json):
    periods = []
    if not resp_json or not isinstance(resp_json, dict):
        return periods
    data = resp_json.get("data")
    data_root = None
    if isinstance(data, list) and data:
        data_root = data[0]
    elif isinstance(data, dict):
        data_root = data
    else:
        data_root = resp_json
    candidate_list_keys = [
        "simples_nacional_periodos_anteriores",
        "simples_nacional_periodos",
        "periodos_simples",
        "simples_periodos",
        "simples_nacional",
        "periodos",
        "permanencia",
        "periodo"
    ]
    for k in candidate_list_keys:
        lst = data_root.get(k) if isinstance(data_root, dict) else None
        if isinstance(lst, list):
            for item in lst:
                if not isinstance(item, dict):
                    continue
                s = _get_value(item, ["inicio_data", "data_inicio", "inicio", "data"])
                e = _get_value(item, ["fim_data", "data_fim", "fim", "data_fim"])
                detalhe = _get_value(item, ["detalhamento", "detalhe", "motivo"]) or ""
                si = parse_date_any(s)
                ei = parse_date_any(e)
                if si:
                    periods.append({"start": si, "end": ei, "detalhe": detalhe})
            if periods:
                return periods
    return periods

def month_date_range(year, month):
    first_day = date(year, month, 1)
    last_day = date(year, month, calendar.monthrange(year, month)[1])
    return first_day, last_day

def is_month_fully_covered(periods, year, month):
    start_month, end_month = month_date_range(year, month)
    hoje = date.today()
    for p in periods:
        si = p.get("start")
        ei = p.get("end")
        detalhe = p.get("detalhe") or ""
        if not si:
            continue
        ei_efetivo = ei if ei else hoje
        if si <= start_month and ei_efetivo >= end_month:
            if year == hoje.year and month == hoje.month and ei is None:
                return True, "Status atual é Simples Nacional."
            else:
                return True, "Permaneceu no Simples Nacional o mês inteiro."
    for p in periods:
        si = p.get("start")
        ei = p.get("end")
        if not si or not ei:
            continue
        if si <= end_month and ei < end_month:
            excl_data = ei.strftime("%Y-%m-%d")
            return False, f"Excluída do Simples Nacional em {excl_data}."
    return False, "Não optante/Nunca esteve no Simples Nacional neste mês."

# ---------------- View Django com download ---------------- #
from django.contrib.auth.decorators import login_required

@login_required(login_url="login")
def upload_file(request):
    if request.method == "POST" and request.FILES.get("file"):
        file = request.FILES["file"]

        try:
            # --- Lê todas as abas do arquivo original ---
            try:
                wb = load_workbook(file)
            except Exception:
                wb = Workbook()  # caso não tenha abas visíveis

            all_sheets = {}
            for sheet_name in wb.sheetnames:
                df_sheet = pd.read_excel(file, sheet_name=sheet_name)
                all_sheets[sheet_name] = df_sheet

            # --- Processa os CNPJs para criar aba "CONSULTA" ---
            cnpjs = read_cnpjs(file)
            rows = []
            start_year = 2020
            hoje = date.today()

            for cnpj in cnpjs:
                status, resp_json = query_infosimples(cnpj)
                periods = extract_periods_from_response(resp_json)

                situacao_atual = None
                if resp_json and "data" in resp_json:
                    data_field = resp_json["data"]
                    if isinstance(data_field, list) and len(data_field) > 0:
                        data_item = data_field[0]
                    elif isinstance(data_field, dict):
                        data_item = data_field
                    else:
                        data_item = {}

                    situacao_atual = (
                        data_item.get("simples_nacional_situacao")
                        or data_item.get("situacao_simples")
                        or data_item.get("situacao")
                    )

                texto_situacao = (situacao_atual or "").lower()
                if (
                        "optante pelo simples nacional" in texto_situacao
                        and "não optante" not in texto_situacao
                        and "nao optante" not in texto_situacao  # cobre casos sem acento
                ):
                    m = re.search(r"desde\s+(\d{2}/\d{2}/\d{4})", texto_situacao)
                    if m:
                        start_date = parse_date_any(m.group(1))
                    else:
                        start_date = date(hoje.year, 1, 1)
                    has_open_period = any(p.get("end") is None for p in periods)
                    if not has_open_period:
                        periods.append({
                            "start": start_date,
                            "end": None,
                            "detalhe": "Situação Atual: Optante pelo Simples Nacional"
                        })

                for year in range(start_year, hoje.year + 1):
                    for month in range(1, 13):
                        if year == hoje.year and month > hoje.month:
                            continue
                        regime, motivo = is_month_fully_covered(periods, year, month)
                        regime_str = "Simples Nacional" if regime else "Outro Regime"
                        mes_data = date(year, month, 1)
                        mes_str = mes_data.strftime("%d/%m/%Y")
                        periods_str = "; ".join([
                            f"{p['start']} - {p.get('end', 'até hoje')} [{p.get('detalhe', '')}]"
                            for p in periods
                        ])
                        rows.append({
                            "CNPJ": cnpj,
                            "MÊS": mes_str,
                            "REGIME": regime_str,
                            "MOTIVO": motivo,
                            "Períodos_detectados": periods_str,
                            "Situacao_Atual": situacao_atual or ""
                        })

                time.sleep(SLEEP)

            df_consulta = pd.DataFrame(rows, columns=["CNPJ", "MÊS", "REGIME", "MOTIVO", "Períodos_detectados", "Situacao_Atual"])

            # --- Salva todas as abas originais + aba "CONSULTA" ---
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                # Reescreve todas as abas originais
                for name, df_sheet in all_sheets.items():
                    df_sheet.to_excel(writer, sheet_name=name, index=False)
                # Adiciona aba de consulta
                df_consulta.to_excel(writer, sheet_name="CONSULTA", index=False)
            output.seek(0)

            response = HttpResponse(
                output.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            response['Content-Disposition'] = f'attachment; filename="consulta_simples.xlsx"'
            return response

        except Exception as e:
            return render(request, "upload.html", {"mensagem": f"❌ Erro ao processar: {e}"})

    return render(request, "upload.html")
